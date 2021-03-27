#!/usr/bin/python3
import cgi
import random
import sys
import ipaddress
import xlrd
import xlwt
from xlwt import Workbook
import openpyxl as op
import smtplib, ssl
# from get_ip_test import *

print("content-type: text/html")
print()


form = cgi.FieldStorage()

cec_name = form.getvalue("cecid")
IP_addr = form.getvalue("ip_extend")
extenddate = form.getvalue("extenddate")
# print(cec_name)
# print(IP_addr)
# print(extenddate)

fileLocation = "available.xlsx"
wb = xlrd.open_workbook(fileLocation) 
sheet = wb.sheet_by_index(0) 
data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]

def mail_sender(data,send_to):
    lst=data.split("_")
    res_dict={}
    res_dict['IP']=lst[1]
    res_dict['Subnet']=lst[2]
    res_dict['Gateway']=lst[3]
    res_dict['Date_From']=lst[4]
    res_dict['Date_To']=lst[5]
    res_dict['Cecid']=lst[6]
    res_dict['Name']=lst[7]

    data_out="\nCongratulations! IP Extend Date Succesful! \n\n IP Allocation Information: \n\n IP Address: "+lst[1]+"\n Subnet: "+lst[2]+"\n Gateway: "+lst[3]+"\n Allocated from: "+lst[4]+"\n Allocated till: "+lst[5]+"\n Allocated to: "+lst[7]+"("+lst[6]+")\n\n Regards,\n Sparsha Lab"    

    #print("inside mailer")
    SERVER = "173.37.102.6"
    FROM = "no-reply@cisco.com"
    rec = str(send_to)+"@cisco.com"
    TO = [rec] # must be a list

    SUBJECT = "IP_Information"
    #TEXT = "You shouldn't have stolen the chocolate. Give it back to him"

    # Prepare actual message
    message = """\
    From: %s
    To: %s
    Subject: "IP Information - Sparsha Lab"

    %s
    """ % (FROM, ", ".join(TO), data_out)
    # Send the mail

    server = smtplib.SMTP(SERVER)
    server.sendmail(FROM, rec, message)
    server.quit()
    #print("email sent")


def extend_date(query):
    #query={"ip":"192.168.1.7","cecid":"robhati","to_date":"6-8-8"}  
    wb_obj=op.load_workbook(fileLocation)
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    flag=0
    result=""
    # result=str(query['ip'])+"_"+str(query['cecid'])
    for r in range (1,max_row+1):
        # result=result+"_"+str(r)+"_"+str(sheet_obj.cell(row=r,column=2).value)+"_"+str(sheet_obj.cell(row=r,column=8).value)
        if (sheet_obj.cell(row=r,column=2).value==str(query['ip']) and sheet_obj.cell(row=r,column=8).value==str(query['cecid'])):
            flag = 1
            # result['ip']=(sheet_obj.cell(row=r,column=2).value)
            # result['mask']=(sheet_obj.cell(row=r,column=3).value)
            # result['gateway']=(sheet_obj.cell(row=r,column=4).value)
            # result['from_date']=(sheet_obj.cell(row=r,column=6).value)
            # result['to_date']=query["to_date"]
            # result['cecid']=(sheet_obj.cell(row=r,column=8).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=2).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=3).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=4).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=6).value)
            result=result+"_"+str(query["to_date"])
            result=result+"_"+str(query["cecid"])
            result=result+"_"+str(sheet_obj.cell(row=r,column=10).value)
            sheet_obj.cell(row=r,column=7).value=query["to_date"]
            # print(result)
            break
    
    
    
    #print(result)
    if flag ==1:
        # print("extended successfully")
       
        # data = '''Hi, \n your {} extended till {}'''.format(query["ip"],query["to_date"])
        mail_sender(result,cec_name)
        print(result)
    else: 
        # print(result)
        print("invalid_inputs")
    wb_obj.save(fileLocation)

query={"ip":IP_addr,"cecid":cec_name,"to_date":extenddate}
extend_date(query)
