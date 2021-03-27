#!/usr/bin/python3
import cgi
import random
import sys
import ipaddress
import xlrd
import xlwt
from xlwt import Workbook 
import openpyxl as op
import smtplib, ssl
import json
import requests 


print("content-type: text/html")
print("")

form = cgi.FieldStorage()

name = form.getvalue("name")
cecid = form.getvalue("cecid")
purpose = form.getvalue("purpose")
fromdate = form.getvalue("fromdate")
todate = form.getvalue("todate")

fileLocation = "available.xlsx"
wb = xlrd.open_workbook(fileLocation) 
sheet = wb.sheet_by_index(0) 
data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]


def getUnallocatedList(data):    
    unallocated_list=[]
    for r in data:
        if r[4]=="unallocated":
            unallocated_list.append(r[1])
    #print(unallocated_list)
    # print(unallocated_list)
    return unallocated_list

def provision_ip(query):
    #query={"ip":"1.1.1.1","from_date":"1-0-2020","to_date":"3-5-2020","cecid":"robhati","purpose":"hjhjs","description":"dsjdsdm"}
    wb_obj=op.load_workbook(fileLocation)
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    result=""
    for r in range (1,max_row+1):
        if (sheet_obj.cell(row=r,column=2).value==query['ip']):
            result=result+"_"+str(sheet_obj.cell(row=r,column=2).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=3).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=4).value)
            result=result+"_"+str(query["from_date"])
            result=result+"_"+str(query["to_date"])
            result=result+"_"+str(query["cecid"])
            result=result+"_"+str(query["name"])
            # print(result)
            break
    
    sheet_obj.cell(row=r,column=5).value="allocated" 
    sheet_obj.cell(row=r,column=6).value=query["from_date"]
    sheet_obj.cell(row=r,column=10).value=query["name"]
    sheet_obj.cell(row=r,column=7).value=query["to_date"]
    sheet_obj.cell(row=r,column=8).value=query["cecid"]
    sheet_obj.cell(row=r,column=9).value=query["purpose"]
    #sheet_obj.cell(row=3,column=4).value="allocated"
    wb_obj.save(fileLocation)
    #print(result)
    # print("Allocation success!!")
    return result

def mail_sender(data,send_to):
    lst=data.split("_")
    res_dict={}
    res_dict['IP']=lst[1]
    res_dict['Subnet']=lst[2]
    res_dict['Gateway']=lst[3]
    res_dict['Date_From']=lst[4]
    res_dict['Date_To']=lst[5]
    res_dict['Cecid']=lst[6]
    res_dict['Name']=lst[7]

    data_out="\nCongratulations! IP Allocation Succesful! \n\n IP Allocation Information: \n\n IP Address: "+lst[1]+"\n Subnet: "+lst[2]+"\n Gateway: "+lst[3]+"\n Allocated from: "+lst[4]+"\n Allocated till: "+lst[5]+"\n Allocated to: "+lst[7]+"("+lst[6]+")\n\n Regards,\n Sparsha Lab"    

    #print("inside mailer")
    SERVER = "173.37.102.6"
    FROM = "no-reply@cisco.com"
    rec = str(send_to)+"@cisco.com"
    TO = [rec] # must be a list

    SUBJECT = "IP_Information"
    #TEXT = "You shouldn't have stolen the chocolate. Give it back to him"

    # Prepare actual message
    message = """\
    From: %s
    To: %s
    Subject: "IP Information - Sparsha Lab"

    %s
    """ % (FROM, ", ".join(TO), data_out)
    # Send the mail

    server = smtplib.SMTP(SERVER)
    server.sendmail(FROM, rec, message)
    server.quit()
    #print("email sent")

if len(getUnallocatedList(data)) >= 1:
    #print(getUnallocatedList(data))
    query={"ip":getUnallocatedList(data)[0],"from_date":fromdate,"to_date":todate,"cecid":cecid,"purpose":purpose,"name":name}
    #query={"ip":getUnallocatedList(data)[0],"from_date":"1-0-2020","to_date":"3-5-2020","cecid":"raghavsi","name":"rohit","purpose":"hjhjs","description":"dsjdsdm"}
    # print("query: ",query)    
    result=provision_ip(query)
    mail_sender(result , cecid)
    # print("E-mail sent")
    print (result)

else:
    print("no_ip_available")


