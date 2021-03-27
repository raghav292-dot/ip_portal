#!/usr/bin/python3
import cgi
import random
import sys
import ipaddress
import xlrd
import xlwt
from xlwt import Workbook 
import openpyxl as op
import smtplib, ssl

print("content-type: text/html")
print()


form = cgi.FieldStorage()

id = form.getvalue("name")
cecid = form.getvalue("cecid")
purpose = form.getvalue("purpose")
description = form.getvalue("description")
ip = form.getvalue("ip")
fromdate = form.getvalue("fromdate")
todate = form.getvalue("todate")
cec_name = form.getvalue("firstname")
IP_addr = form.getvalue("ip_addr")


#print(id)
#print(cecid)
#print(purpose)
#print(description)
#print(ip)
#print(fromdate)
#print(todate)

fileLocation = "available.xlsx"
wb = xlrd.open_workbook(fileLocation) 
sheet = wb.sheet_by_index(0) 
data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]


def getUnallocatedList(data):    
    unallocated_list=[]
    for r in data:
        if r[4]=="unallocated":
            unallocated_list.append(r[1])
    #print(unallocated_list)
    return unallocated_list

def provision_ip(query):
    #query={"ip":"1.1.1.1","from_date":"1-0-2020","to_date":"3-5-2020","cecid":"robhati","purpose":"hjhjs","description":"dsjdsdm"}
    wb_obj=op.load_workbook(fileLocation)
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    result={}
    for r in range (1,max_row):
        if (sheet_obj.cell(row=r,column=2).value==query['ip']):
            result['ip']=(sheet_obj.cell(row=r,column=2).value)
            result['mask']=(sheet_obj.cell(row=r,column=3).value)
            result['gateway']=(sheet_obj.cell(row=r,column=4).value)
            result['from_date']=query["from_date"]
            result['to_date']=query["to_date"]
            result['cecid']=query["cecid"]
            #print(result)
            break
    
    sheet_obj.cell(row=r,column=5).value="allocated" 
    sheet_obj.cell(row=r,column=6).value=query["from_date"]
    sheet_obj.cell(row=r,column=7).value=query["to_date"]
    sheet_obj.cell(row=r,column=8).value=query["cecid"]
    sheet_obj.cell(row=r,column=9).value=query["purpose"]
    sheet_obj.cell(row=r,column=10).value=query["description"]
    #sheet_obj.cell(row=3,column=4).value="allocated"
    wb_obj.save(fileLocation)
    print(result)
    print("allocation success")
    return result;

def mail_sender(data):
    #print("inside mailer")
    SERVER = "173.37.102.6"
    FROM = "no-reply@cisco.com"
    TO = ["raghavsi@cisco.com"] # must be a list

    SUBJECT = "IP_Information"
    #TEXT = "You shouldn't have stolen the chocolate. Give it back to him"

    # Prepare actual message
    message = """\
    From: %s
    To: %s
    Subject: %s

    %s
    """ % (FROM, ", ".join(TO), SUBJECT, data)
    # Send the mail

    server = smtplib.SMTP(SERVER)
    server.sendmail(FROM, TO, message)
    server.quit()



if len(getUnallocatedList(data)) > 1:
    #print(getUnallocatedList(data)[0])
    query={"ip":getUnallocatedList(data)[0],"from_date":fromdate,"to_date":todate,"cecid":cecid,"purpose":purpose,"description":description}
    print("result: ")    
    provision_ip(query)
#    mail_sender(provision_ip(query))
    print("E-mail sent")
else:
    print("result: "+"no ip available")


