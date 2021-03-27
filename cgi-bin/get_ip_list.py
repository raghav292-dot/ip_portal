#!/usr/bin/python3
import cgi
import random
import sys
import ipaddress
import xlrd
import xlwt
from xlwt import Workbook 
import openpyxl as op
import smtplib, ssl
import json
import requests 


print("content-type: text/html")
print("")

# form = cgi.FieldStorage()

# name = form.getvalue("name")
# cecid = form.getvalue("cecid")
# purpose = form.getvalue("purpose")
# fromdate = form.getvalue("fromdate")
# todate = form.getvalue("todate")

fileLocation = "available.xlsx"
wb = xlrd.open_workbook(fileLocation) 
sheet = wb.sheet_by_index(0) 
data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]


def getUnallocatedList(data):    
    unallocated_list=[]
    for r in data:
        if r[4]=="unallocated":
            unallocated_list.append(r[1])
    #print(unallocated_list)
    # print(unallocated_list)
    return unallocated_list

def provision_ip(query):
    #query={"ip":"1.1.1.1","from_date":"1-0-2020","to_date":"3-5-2020","cecid":"robhati","purpose":"hjhjs","description":"dsjdsdm"}
    wb_obj=op.load_workbook(fileLocation)
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    result=""
    for r in range (1,max_row+1):
        if (sheet_obj.cell(row=r,column=2).value==query['ip']):
            result=result+"_"+str(sheet_obj.cell(row=r,column=2).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=3).value)
            result=result+"_"+str(sheet_obj.cell(row=r,column=4).value)
            result=result+"_"+str(query["from_date"])
            result=result+"_"+str(query["to_date"])
            result=result+"_"+str(query["cecid"])
            result=result+"_"+str(query["name"])
            # print(result)
            break
    
    sheet_obj.cell(row=r,column=5).value="allocated" 
    sheet_obj.cell(row=r,column=6).value=query["from_date"]
    sheet_obj.cell(row=r,column=10).value=query["name"]
    sheet_obj.cell(row=r,column=7).value=query["to_date"]
    sheet_obj.cell(row=r,column=8).value=query["cecid"]
    sheet_obj.cell(row=r,column=9).value=query["purpose"]
    #sheet_obj.cell(row=3,column=4).value="allocated"
    wb_obj.save(fileLocation)
    #print(result)
    # print("Allocation success!!")
    return result

#def mail_sender(data):
#    print("inside mailer")
#    print(data)    
#    port = 465  # For SSL
#    smtp_server = "smtp.gmail.com"
#    sender_email = "rs1099867@gmail.com"  # Enter your address
#    receiver_email = "singhraghvendra948@gmail.com"  # Enter receiver address
#   password = "raghvendra1"
#    #data="raghav"
#    message = """\
#    Subject: Ip information

#   This email contains ip information"""
#    context = ssl.create_default_context()
#    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
#        server.login(sender_email, password)
#        server.sendmail(sender_email, receiver_email, message+str(data))



if len(getUnallocatedList(data)) >= 1:
    # print(getUnallocatedList(data)[0])
    # query={"ip":getUnallocatedList(data)[0],"from_date":fromdate,"to_date":todate,"cecid":cecid,"purpose":purpose,"name":name}
    # # print("query: ",query)    
    # result=provision_ip(query)
#    mail_sender(provision_ip(query))
    # print("E-mail sent")
    print ("complete")

else:
    print("no_ip_available")




# #!/usr/bin/python3
# import cgi
# import random
# import sys
# import ipaddress
# import xlrd
# import xlwt
# from xlwt import Workbook 
# import openpyxl as op
# import smtplib, ssl
# import json
# import requests 


# print("content-type: text/html")
# print("")
# print("hello")

# # form = cgi.FieldStorage()
# # cecid = form.getvalue("firstname")


# # fileLocation = "available.xlsx"
# # wb = xlrd.open_workbook(fileLocation) 
# # sheet = wb.sheet_by_index(0) 
# # data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]


# # def getallocatedList(data,cecid):    
# #     result=str(cecid)
# #     for r in data:
# #         if r[4]=="allocated"and r[7]==cecid:
# #              result=result+"_"+str(r[1])
           
# #     #print(unallocated_list)
# #     # print(unallocated_list)
# #     return result

# # print(getallocatedList(data,cecid))


