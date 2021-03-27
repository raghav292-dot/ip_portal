#!/usr/bin/python3
import cgi
import random
import sys
import ipaddress
import xlrd
import xlwt
from xlwt import Workbook
import openpyxl as op
import smtplib, ssl

print("content-type: text/html")
print()



form = cgi.FieldStorage()

cec_name = form.getvalue("cecid")
IP_addr = form.getvalue("ip_return")
#print(cec_name)
#print(IP_addr)
# print(cec_name,IP_addr)

fileLocation = "available.xlsx"
wb = xlrd.open_workbook(fileLocation) 
sheet = wb.sheet_by_index(0) 
data=[[sheet.cell_value(r,c) for c in range (sheet.ncols)]for r in range (sheet.nrows)]


def mail_sender():
    data_out="\nCongratulations! IP Unallocation Succesful! \n\n IP Unallocation Information: \n\n IP Address: "+IP_addr+"\n Unallocated to: "+cec_name+")\n\n Regards,\n Sparsha Lab"    

    #print("inside mailer")
    SERVER = "173.37.102.6"
    FROM = "no-reply@cisco.com"
    rec = str(cec_name)+"@cisco.com"
    TO = [rec] # must be a list

    SUBJECT = "IP_Information"
    #TEXT = "You shouldn't have stolen the chocolate. Give it back to him"

    # Prepare actual message
    message = """\
    From: %s
    To: %s
    Subject: "IP Information - Sparsha Lab"

    %s
    """ % (FROM, ", ".join(TO), data_out)
    # Send the mail

    server = smtplib.SMTP(SERVER)
    server.sendmail(FROM, rec, message)
    server.quit()
    #print("email sent")

def unallocate_ip(query):
    #query={"ip":"192.168.1.5","cecid":"robhati"}   
    wb_obj=op.load_workbook(fileLocation)
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    result={}
    ip_flag=0
    
    for r in range (1,max_row+1):
            if (sheet_obj.cell(row=r,column=2).value==query['ip'] and sheet_obj.cell(row=r,column=8).value==query['cecid']):
                ip_flag = 1
                result['ip']=(sheet_obj.cell(row=r,column=2).value)
                result['mask']=(sheet_obj.cell(row=r,column=3).value)
                result['gateway']=(sheet_obj.cell(row=r,column=4).value)
                result['from_date']=(sheet_obj.cell(row=r,column=6).value)
                result['to_date']=(sheet_obj.cell(row=r,column=7).value)
                result['cecid']=(sheet_obj.cell(row=r,column=8).value)
                #print(result)
                
                sheet_obj.cell(row=r,column=5).value="unallocated" 
                sheet_obj.cell(row=r,column=6).value=""
                sheet_obj.cell(row=r,column=7).value=""
                sheet_obj.cell(row=r,column=8).value=""
                sheet_obj.cell(row=r,column=9).value=""
                sheet_obj.cell(row=r,column=10).value=""
                wb_obj.save(fileLocation)
                # print("unallocation success")
                break
        
        
    
    if ip_flag==0:
        print("invalid_inputs")
    else:
        mail_sender()
        print("unallocation success")


query={"ip":IP_addr,"cecid":cec_name}
unallocate_ip(query)

